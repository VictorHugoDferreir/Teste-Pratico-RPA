from pathlib import Path
import openpyxl

from modelos.cliente import Cliente
from configuracao import configuracoes


def ler_planilha(caminho_arquivo: Path) -> list[Cliente]:
    print("[INFO] Lendo planilha...")

    # Carrega a planilha usando openpyxl, com data_only=True para obter os valores das células em vez das fórmulas.
    planilha = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    aba = planilha.active

    cabecalhos = {
        nome: indice
        for indice, nome in enumerate(
            next(aba.iter_rows(max_row=1, values_only=True))
        )
        if nome
    } #cria um dicionário que mapeia os nomes dos cabeçalhos para seus índices de coluna correspondentes na primeira linha da planilha.

    # Verifica se todos os cabeçalhos obrigatórios estão presentes na planilha.
    for coluna in configuracoes.CABECALHOS_OBRIGATORIOS:
        if coluna not in cabecalhos:
            raise ValueError(
                f"Cabecalho '{coluna}' nao encontrado na planilha."
            )

    clientes = []

    #para cada linha da planilha, a partir da segunda linha, obtem apenas os valores das células.
    for linha in aba.iter_rows(min_row=2, values_only=True):

        if not any(linha):
            continue
        
        #Lista de objetos Cliente, cada um representando uma linha da planilha, com os campos preenchidos a partir dos valores das células correspondentes.
        cliente = Cliente(
            email_solicitante=str(
                linha[cabecalhos["Email do Solicitante"]] or ""
            ).strip(),

            montante_emprestimo=str(
                linha[cabecalhos["Montante do Empréstimo"]] or ""
            ).strip(),

            termo_emprestimo=str(
                int(float(linha[cabecalhos["Termo do Empréstimo"]]))
            ).strip(),

            renda_anual_atual=str(
                linha[cabecalhos["Renda Anual Atual( Antes dos Impostos)"]] or ""
            ).strip(),

            idade=str(
                linha[cabecalhos["Idade"]] or ""
            ).strip(),
        )

        clientes.append(cliente)

    print(f"[INFO] {len(clientes)} cliente(s) encontrado(s).")

    return clientes